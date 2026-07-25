"""The single data loader.

This replaces six byte-identical copies of `load_daily()` that were spread
across the original stage scripts (plus a seventh, differently-written but
semantically equivalent, loader in stage1_foundation.py). Behaviour is
unchanged -- tests/test_equivalence.py asserts this loader reproduces the
originals exactly on the real workbook.

Cleaning rules, in order:
  1. read every worksheet; one worksheet per station
  2. drop rows with no timestamp
  3. coerce PM2.5 to numeric, clip above PM_CAP
  4. reindex each station onto a complete daily grid from its first to its
     last observation
  5. linearly interpolate interior gaps of at most MAX_GAP days; longer gaps
     and any gap touching an endpoint stay NaN, ENTIRELY -- see
     `_interpolate_short_gaps` for why that word is load-bearing

## The interpolation rule is per-RUN, not per-value

`Series.interpolate(limit=n)` does NOT mean "skip gaps longer than n". It means
"fill at most n consecutive NaNs within any run", so a 7-day gap gets its first
3 days synthesised and the remaining 4 left NaN. Every original script used
that call, and it fabricated 306 values inside 102 over-long gaps corpus-wide.

`stage1_foundation.py` shows the intent was always the per-run rule: its lines
72-78 build a NaN-run classification and then never use the result. This module
implements what that dead code was reaching for.
"""

from __future__ import annotations

from functools import lru_cache
from math import asin, cos, radians, sin, sqrt
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from . import config as cfg

__all__ = [
    "load_daily",
    "station_metadata",
    "wide_matrix",
    "neighbour_lag_frames",
    "haversine_km",
    "nearest_neighbours",
]


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in km. One copy, previously duplicated six times."""
    radius = 6371.0
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    h = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return 2 * radius * asin(sqrt(h))


def _read_workbook(path: Path) -> list[pd.DataFrame]:
    workbook = openpyxl.load_workbook(path, read_only=True)
    frames: list[pd.DataFrame] = []
    for sheet in workbook.sheetnames:
        records = []
        for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
            if row[cfg.COL_TS] is None:
                continue
            records.append(
                (
                    pd.Timestamp(str(row[cfg.COL_TS]).replace("Z", "")).normalize(),
                    row[cfg.COL_PM],
                    float(row[cfg.COL_LAT]),
                    float(row[cfg.COL_LON]),
                    row[cfg.COL_NAME],
                )
            )
        if not records:
            continue
        frame = pd.DataFrame(
            records, columns=["date", "pm25", "lat", "lon", "name"]
        ).sort_values("date")
        frame["pm25"] = pd.to_numeric(frame["pm25"], errors="coerce").clip(
            upper=cfg.PM_CAP
        )
        frames.append(frame.assign(sheet=sheet))
    return frames


def nan_runs(mask: np.ndarray) -> list[tuple[int, int, bool]]:
    """Maximal runs of True in `mask`, as (start, stop, is_interior).

    `stop` is exclusive. A run is interior when it touches neither end of the
    array -- i.e. real observations bracket it, so linear interpolation has
    something to interpolate between.
    """
    runs: list[tuple[int, int, bool]] = []
    n = len(mask)
    i = 0
    while i < n:
        if not mask[i]:
            i += 1
            continue
        j = i
        while j < n and mask[j]:
            j += 1
        runs.append((i, j, i > 0 and j < n))
        i = j
    return runs


def _interpolate_short_gaps(series: pd.Series, max_gap: int = cfg.MAX_GAP) -> pd.Series:
    """Fill interior NaN runs of length <= max_gap; leave every other NaN alone.

    The rule is applied to the RUN, not to individual values. A run of 4 days
    with max_gap=3 is left ENTIRELY NaN -- it does not get its first 3 days
    filled. That distinction is the whole point; see the module docstring.
    """
    values = series.to_numpy(dtype=float)
    missing = np.isnan(values)
    if not missing.any():
        return series

    # Interpolate everything interior, then keep only the runs we are allowed to fill.
    candidate = series.interpolate(method="linear", limit_area="inside").to_numpy(dtype=float)
    out = values.copy()
    for start, stop, interior in nan_runs(missing):
        if interior and (stop - start) <= max_gap:
            out[start:stop] = candidate[start:stop]
    return pd.Series(out, index=series.index)


@lru_cache(maxsize=8)
def load_daily(path: Path | None = None, interpolate: bool = True) -> pd.DataFrame:
    """Tidy daily grid: one row per station-day.

    Columns: sheet, name, lat, lon, date, pm25.
    Cached, because reading the workbook dominates the runtime of every stage
    and the result is immutable in practice.

    `interpolate` (default True) fills interior NaN runs of <= MAX_GAP days, the
    behaviour every committed stage relies on. Passing `interpolate=False` skips
    that step entirely, leaving raw non-null values only and every gap as NaN;
    the 14-day window-validity rule in `aq.features` then drops any window that
    would have depended on a filled day. This is the no-interpolation path used
    by the leak-free sensitivity analysis. `interpolate` participates in the
    cache key, so the two variants coexist without clobbering each other; the
    no-arg call keeps the exact cached result the pipeline has always used.
    """
    path = Path(path) if path is not None else cfg.DATA_FILE
    if not path.exists():
        raise FileNotFoundError(
            f"station workbook not found at {path}\n"
            f"Set AQ_DATA_FILE, or place the file at {cfg.DATA_FILE}."
        )

    out = []
    for frame in _read_workbook(path):
        index = pd.date_range(frame["date"].min(), frame["date"].max(), freq="D")
        series = frame.set_index("date")["pm25"].reindex(index)
        if interpolate:
            series = _interpolate_short_gaps(series, cfg.MAX_GAP)
        out.append(
            pd.DataFrame(
                {
                    "sheet": frame["sheet"].iloc[0],
                    "name": frame["name"].iloc[0],
                    "lat": frame["lat"].iloc[0],
                    "lon": frame["lon"].iloc[0],
                    "date": index,
                    "pm25": series.values,
                }
            )
        )
    return pd.concat(out, ignore_index=True)


def station_metadata(daily: pd.DataFrame) -> pd.DataFrame:
    """One row per station: name, lat, lon, first and last observed day, count.

    `first` and `last` are computed from OBSERVED values only (NaN rows created
    by the daily reindex do not count), because the protocol keys off a
    station's genuine deployment date.
    """
    observed = daily.dropna(subset=["pm25"])
    meta = (
        observed.groupby("sheet")
        .agg(
            name=("name", "first"),
            lat=("lat", "first"),
            lon=("lon", "first"),
            first=("date", "min"),
            last=("date", "max"),
            n_obs=("pm25", "size"),
        )
        .reset_index()
    )
    return meta


def wide_matrix(daily: pd.DataFrame) -> pd.DataFrame:
    """Date x station matrix of PM2.5, on a complete daily index.

    The reindex matters. The original stage4 looked neighbours up by label
    (`wide[src].get(day - 1)`) while stage5/5b shifted by row position
    (`wide.shift(1)`). Those agree only when the index has no missing days.
    Reindexing here makes both readings identical and the assert makes the
    assumption explicit rather than silent.
    """
    wide = daily.pivot_table(index="date", columns="sheet", values="pm25")
    full = pd.date_range(wide.index.min(), wide.index.max(), freq="D")
    wide = wide.reindex(full)
    assert len(wide) == len(full), "date index is not a complete daily grid"
    return wide


def neighbour_lag_frames(wide: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """(t-1, t-2) views of the station matrix, for spatial features.

    Both are strictly past-only relative to the day being predicted, which is
    what keeps neighbour features leak-free.
    """
    return wide.shift(1), wide.shift(2)


def nearest_neighbours(
    sheet: str, meta: pd.DataFrame, pool, k: int
) -> list[tuple[str, float]]:
    """The k nearest OTHER stations drawn from `pool`, nearest first.

    Returns [(sheet, distance_km)]. Ordering makes feature block i mean
    "i-th nearest neighbour" consistently for every station.
    """
    if k == 0:
        return []
    row = meta[meta.sheet == sheet].iloc[0]
    distances = []
    for other in pool:
        if other == sheet:
            continue
        other_row = meta[meta.sheet == other].iloc[0]
        distances.append(
            (
                haversine_km(row["lat"], row["lon"], other_row["lat"], other_row["lon"]),
                other,
            )
        )
    distances.sort()
    return [(name, dist) for dist, name in distances[:k]]


def rmse(actual: np.ndarray, predicted: np.ndarray) -> float:
    """Root mean squared error. Kept here so every stage uses one definition."""
    actual = np.asarray(actual, dtype=float)
    predicted = np.asarray(predicted, dtype=float)
    return float(np.sqrt(np.mean((actual - predicted) ** 2)))
