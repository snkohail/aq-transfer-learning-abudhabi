#!/usr/bin/env python3
"""Analysis stages.

Each `stage_*` function is a self-contained analysis: it builds its designs,
scores every method on a single common index (via aq.evaluate), writes one CSV
to `out_dir`, and returns the DataFrame. `run.py` is only the CLI that dispatches
to these; keeping them here makes them importable and testable without the
argument parser (see tests/test_regression.py).

The one rule these all obey: never derive an evaluation mask from a single
design. Build it once with build_common_index() across every arm, then score.
"""

from __future__ import annotations

from datetime import timedelta

import numpy as np
import pandas as pd
from sklearn.linear_model import Ridge, RidgeCV
from sklearn.metrics import f1_score, precision_score, recall_score
from sklearn.preprocessing import StandardScaler

from aq import config as cfg
from aq.data import (
    load_daily,
    nearest_neighbours,
    neighbour_lag_frames,
    station_metadata,
    wide_matrix,
)
from aq.evaluate import Prediction, build_common_index, align, score, rmse
from aq.features import build_design, build_sequences
from aq.models import RidgeBackbone
from aq.splits import build_manifest, select_targets, source_frames, train_end_for
from aq.stats import compare, summarise


def log(*args):
    print(*args, flush=True)


def banner(title: str):
    log("\n" + "=" * 78)
    log(title)
    log("=" * 78)


def _raw_record_counts() -> dict:
    """Row counts straight from the workbook, before any reindex or interpolation.

    The paper must cite the RAW non-null count as its sample size, not the
    workbook row count and not the post-interpolation count.
    """
    import openpyxl

    rows = non_null = 0
    workbook = openpyxl.load_workbook(cfg.DATA_FILE, read_only=True)
    for sheet in workbook.sheetnames:
        for row in workbook[sheet].iter_rows(min_row=2, values_only=True):
            if row[cfg.COL_TS] is None:
                continue
            rows += 1
            if pd.notna(pd.to_numeric(row[cfg.COL_PM], errors="coerce")):
                non_null += 1
    return {"rows": rows, "non_null": non_null}


def _context():
    daily = load_daily()
    meta = station_metadata(daily)
    wide = wide_matrix(daily)
    wide1, wide2 = neighbour_lag_frames(wide)
    return daily, meta, wide1, wide2


def _pooled_source_design(daily, meta, train_end, target_first, k_neighbours, wide1, wide2):
    """Stack every eligible source into one training matrix."""
    frames = source_frames(daily, train_end, target_first)
    blocks_X, blocks_y = [], []
    for sheet, frame in frames.items():
        neighbours = nearest_neighbours(sheet, meta, cfg.RICH_SOURCES, k_neighbours)
        if len(neighbours) != k_neighbours:
            continue
        design = build_design(
            frame,
            neighbours=neighbours or None,
            wide1=wide1 if neighbours else None,
            wide2=wide2 if neighbours else None,
        )
        if len(design) > cfg.MIN_SOURCE_ROWS:
            blocks_X.append(design.X)
            blocks_y.append(design.y)
    if not blocks_X or len({b.shape[1] for b in blocks_X}) != 1:
        return None, None
    return np.vstack(blocks_X), np.concatenate(blocks_y)


# ---------------------------------------------------------------------------
# stages
# ---------------------------------------------------------------------------
def stage_audit(out_dir):
    daily, meta, _, _ = _context()
    banner("AUDIT — corpus facts (FINDINGS.md Section 2)")

    # Report the raw and post-interpolation counts separately. Conflating them
    # is how "28,806" ended up cited as the sample size in the old drafts.
    raw = _raw_record_counts()
    observed = daily["pm25"].dropna()
    log(f"stations                      {daily.sheet.nunique()}")
    log(f"raw rows in the workbook      {raw['rows']}")
    log(f"raw non-null PM2.5            {raw['non_null']}   <- the sample size to cite")
    log(f"rows on the daily grid        {len(daily)}")
    log(f"non-null after interpolation  {len(observed)}")
    log(f"range                         {observed.min():.2f} - {observed.max():.2f} ug/m3")
    log(f"median / mean                 {observed.median():.1f} / {observed.mean():.1f}")
    log(f"date span                     {daily.date.min().date()} -> {daily.date.max().date()}")
    log(f"days >= {cfg.EXTREME_ABS:.0f} ug/m3       "
        f"{(observed >= cfg.EXTREME_ABS).sum()} "
        f"({100 * (observed >= cfg.EXTREME_ABS).mean():.1f}%)")

    targets = select_targets(daily, meta)
    log(f"\nrich sources ({len(cfg.RICH_SOURCES)}): {', '.join(cfg.RICH_SOURCES)}")
    log(f"authentic targets ({len(targets)}): {', '.join(targets)}")

    distances = []
    for sheet in targets:
        nearest = nearest_neighbours(sheet, meta, cfg.RICH_SOURCES, 1)
        if nearest:
            distances.append(nearest[0][1])
    if distances:
        log(f"median nearest-source distance {np.median(distances):.1f} km "
            f"(range {min(distances):.1f}-{max(distances):.1f})")

    manifest = build_manifest(daily, meta)
    path = out_dir / "manifest.csv"
    manifest.to_csv(path, index=False)
    log(f"\nsaved -> {path}")
    return manifest


def stage_selection(out_dir):
    """Does choosing sources beat pooling all of them? (FINDINGS 5.2: no.)"""
    daily, meta, wide1, wide2 = _context()
    targets = select_targets(daily, meta)
    banner("SELECTION — source choice vs uniform pooling")

    rows = []
    for sheet in targets:
        row = meta[meta.sheet == sheet].iloc[0]
        train_end = train_end_for(row["first"])
        frames = source_frames(daily, train_end, row["first"])

        parts, distances = {}, {}
        for source, frame in frames.items():
            design = build_design(frame)
            if len(design) > cfg.MIN_SOURCE_ROWS:
                parts[source] = design
                near = [d for s, d in nearest_neighbours(sheet, meta, [source], 1)]
                distances[source] = near[0] if near else np.inf
        if len(parts) < cfg.MIN_SOURCES:
            continue

        target_design = build_design(daily[daily.sheet == sheet][["date", "pm25"]])
        by_distance = sorted(parts, key=lambda s: distances[s])

        for k_adapt in cfg.K_LIST:
            test_start = row["first"] + timedelta(days=k_adapt)
            index = build_common_index(target_design, not_before=test_start)
            if len(index) == 0:
                continue
            X_test, y_test = align(target_design, index)
            persistence = X_test[:, cfg.SEQ_L - 1]

            record = {
                "target": sheet,
                "name": row["name"],
                "K": k_adapt,
                "n_test": len(index),
                "persistence": rmse(y_test, persistence),
            }

            def fit_predict(selected, weights=None):
                X = np.vstack([parts[s].X for s in selected])
                y = np.concatenate([parts[s].y for s in selected])
                sample_weight = (
                    np.concatenate([np.full(len(parts[s].y), weights[s]) for s in selected])
                    if weights
                    else None
                )
                model = RidgeBackbone().fit_sources(X, y, sample_weight=sample_weight)
                return rmse(y_test, model.predict(X_test))

            record["pooled_all"] = fit_predict(by_distance)
            for k in (3, 5):
                record[f"pooled_top{k}_geo"] = fit_predict(by_distance[:k])
            decay = {s: float(np.exp(-distances[s] / cfg.GEO_TAU)) for s in parts}
            record["pooled_geo_weighted"] = fit_predict(by_distance, decay)

            rng = np.random.default_rng(cfg.SEED)
            for k in (3, 5):
                draws = [
                    fit_predict(list(rng.choice(list(parts), size=k, replace=False)))
                    for _ in range(20)
                ]
                record[f"pooled_rand{k}"] = float(np.mean(draws))
            rows.append(record)

    frame = pd.DataFrame(rows)
    path = out_dir / "selection.csv"
    frame.to_csv(path, index=False)

    for k_adapt in sorted(frame.K.unique()):
        subset = frame[frame.K == k_adapt]
        log(f"\n--- K={k_adapt} (n={len(subset)}) ---")
        comparisons = [
            compare(subset.pooled_all, subset[col], label=col)
            for col in ["pooled_top3_geo", "pooled_top5_geo", "pooled_geo_weighted",
                        "pooled_rand3", "pooled_rand5"]
            if col in subset
        ]
        log(summarise(comparisons))
    log(f"\nsaved -> {path}")
    return frame


def stage_ablation(out_dir):
    """Neighbour ablation on ONE evaluation index shared by every k."""
    daily, meta, wide1, wide2 = _context()
    targets = select_targets(daily, meta)
    k_values = [0, 1, 2, 3, 5, 8]
    banner("ABLATION — neighbour features, common evaluation index")

    rows = []
    for sheet in targets:
        row = meta[meta.sheet == sheet].iloc[0]
        train_end = train_end_for(row["first"])
        test_start = row["first"] + timedelta(days=cfg.K_ADAPT_DEFAULT)
        series = daily[daily.sheet == sheet][["date", "pm25"]]

        designs, neighbour_sets = {}, {}
        for k in k_values:
            neighbours = nearest_neighbours(sheet, meta, cfg.RICH_SOURCES, k)
            neighbour_sets[k] = neighbours
            designs[k] = build_design(
                series,
                neighbours=neighbours or None,
                wide1=wide1 if neighbours else None,
                wide2=wide2 if neighbours else None,
            )

        # THE guard rail: intersect every arm before scoring any of them.
        index = build_common_index(*designs.values(), not_before=test_start)
        if len(index) == 0:
            log(f"skip {row['name'][:34]} (common test days < {cfg.MIN_TEST})")
            continue

        # CAREFUL -- "extreme" means three different things in the original code:
        #   stage4  absolute >=75 ug/m3            -> our `ext` slice
        #   stage4  90th pct of the FULL series    -> our `extrel` slice
        #   stage5b 90th pct of the COMMON TEST y  -> what its `rmse_ext` column held
        # The handoff inherits the ambiguity: Section 4.3's "extreme (>=75)" is
        # absolute, Section 5.5's "RMSE extreme" is stage5b's test-window
        # relative threshold. They are not the same slice despite the same word.
        # We reproduce stage5b here so Section 5.5 is comparable, and report the
        # absolute slice alongside it rather than instead of it.
        _, y_reference = align(designs[0], index)
        threshold = float(np.nanpercentile(y_reference, cfg.EXTREME_PCT))

        for k in k_values:
            X_train, y_train = _pooled_source_design(
                daily, meta, train_end, row["first"], k, wide1, wide2
            )
            if X_train is None:
                continue
            X_test, y_test = align(designs[k], index)
            if X_test.shape[1] != X_train.shape[1]:
                continue
            model = RidgeBackbone().fit_sources(X_train, y_train)

            predictions = {
                "model": Prediction(model.predict(X_test), index),
                "persistence": Prediction(X_test[:, cfg.SEQ_L - 1], index),
            }
            scores = score(
                y_test, predictions, index,
                metrics=("rmse", "mae"), relative_threshold=threshold,
            )
            rows.append({
                "target": sheet, "name": row["name"], "k": k, "n_common": len(index),
                "kth_dist": max([d for _, d in neighbour_sets[k]], default=0.0),
                **scores,
            })

    frame = pd.DataFrame(rows)
    path = out_dir / "ablation.csv"
    frame.to_csv(path, index=False)

    baseline = frame[frame.k == 0].set_index("target")
    log(f"\ntargets={frame.target.nunique()}  median common test days="
        f"{frame.n_common.median():.0f}\n")
    comparisons = []
    for k in k_values:
        if k == 0:
            continue
        arm = frame[frame.k == k].set_index("target")
        shared = baseline.index.intersection(arm.index)
        comparisons.append(
            compare(baseline.loc[shared, "rmse_all_model"],
                    arm.loc[shared, "rmse_all_model"], label=f"k={k} overall")
        )
        comparisons.append(
            compare(baseline.loc[shared, "rmse_extrel_model"],
                    arm.loc[shared, "rmse_extrel_model"], label=f"k={k} extreme")
        )
    log(summarise(comparisons))
    log(f"\nsaved -> {path}")
    return frame


def _detection(y_true, y_pred, threshold):
    """Precision, recall, F1 and event count for next-day exceedance of `threshold`.

    An "event" is an observed day at or above the per-station relative threshold;
    a "warning" is a predicted day at or above it. zero_division=0 so a model that
    never warns scores 0 rather than raising.
    """
    is_event = np.asarray(y_true) >= threshold
    warned = np.asarray(y_pred) >= threshold
    return (
        float(precision_score(is_event, warned, zero_division=0)),
        float(recall_score(is_event, warned, zero_division=0)),
        float(f1_score(is_event, warned, zero_division=0)),
        int(is_event.sum()),
    )


def stage_extreme(out_dir):
    """Where the aggregate gain comes from, and where it does not -- plus the
    event-warning skill for next-day exceedance of the per-station top-10%
    threshold (precision / recall / F1 for ridge and persistence).

    The LSTM's warning skill is reported by stage_lstm, which is the stage that
    trains it; keeping the LSTM out of here keeps this stage (and the test suite)
    torch-free and fast.
    """
    daily, meta, wide1, wide2 = _context()
    targets = select_targets(daily, meta)
    banner("EXTREME — error decomposition by slice + event-warning skill")

    rows, det = [], []
    for sheet in targets:
        row = meta[meta.sheet == sheet].iloc[0]
        train_end = train_end_for(row["first"])
        test_start = row["first"] + timedelta(days=cfg.K_ADAPT_DEFAULT)
        series = daily[daily.sheet == sheet][["date", "pm25"]]

        design = build_design(series)
        index = build_common_index(design, not_before=test_start)
        if len(index) == 0:
            continue

        X_train, y_train = _pooled_source_design(
            daily, meta, train_end, row["first"], 0, wide1, wide2
        )
        if X_train is None:
            continue
        X_test, y_test = align(design, index)
        model = RidgeBackbone().fit_sources(X_train, y_train)

        threshold = float(np.nanpercentile(design.y, cfg.EXTREME_PCT))
        predictions = {
            "model": Prediction(model.predict(X_test), index),
            "persistence": Prediction(X_test[:, cfg.SEQ_L - 1], index),
        }
        scores = score(y_test, predictions, index, relative_threshold=threshold)
        rows.append({"target": sheet, "name": row["name"], "n_test": len(index), **scores})

        # event-warning skill, only where there are enough events to be meaningful
        if (y_test >= threshold).sum() >= cfg.MIN_SLICE_DAYS:
            for name, pred in (("ridge", predictions["model"]),
                               ("persistence", predictions["persistence"])):
                p, r, f, n_ev = _detection(y_test, pred.values, threshold)
                det.append({"target": sheet, "name": row["name"], "model": name,
                            "precision": p, "recall": r, "f1": f, "n_events": n_ev})

    frame = pd.DataFrame(rows)
    frame.to_csv(out_dir / "extreme.csv", index=False)
    det_frame = pd.DataFrame(det)
    det_frame.to_csv(out_dir / "detection.csv", index=False)

    comparisons = [
        compare(frame[f"rmse_{s}_persistence"], frame[f"rmse_{s}_model"], label=s)
        for s in ("all", "normal", "ext", "extrel", "onset")
        if f"rmse_{s}_model" in frame
    ]
    log(summarise(comparisons))
    log("\nSlices with fewer than "
        f"{cfg.MIN_SLICE_DAYS} qualifying days report NaN, which is why n differs "
        "by slice. That is a reporting rule, not an absence of events.")

    log("\n" + "-" * 60)
    log(f"EVENT-WARNING SKILL — next-day exceedance of the top-{100 - cfg.EXTREME_PCT}% "
        f"threshold  (median over {det_frame.target.nunique()} targets)")
    log(f"{'model':<14}{'precision':>11}{'recall':>9}{'F1':>8}")
    for name in ("ridge", "persistence"):
        s = det_frame[det_frame.model == name]
        log(f"{name:<14}{s.precision.median():>11.2f}{s.recall.median():>9.2f}{s.f1.median():>8.2f}")
    ridge_rec = det_frame[det_frame.model == "ridge"].recall.median()
    log(f"-> ridge misses ~{100 * (1 - ridge_rec):.0f}% of events (recall {ridge_rec:.2f})")
    log(f"\nsaved -> {out_dir / 'extreme.csv'}, {out_dir / 'detection.csv'}")
    return frame


def stage_lstm(out_dir):
    """LSTM+attention vs ridge vs persistence. Needs torch."""
    from aq.models import LSTMBackbone
    from sklearn.preprocessing import StandardScaler

    daily, meta, _, _ = _context()
    targets = select_targets(daily, meta)
    banner("LSTM — recurrent backbone vs ridge vs persistence")

    rows, det = [], []
    cache = {}
    for n, sheet in enumerate(targets, 1):
        row = meta[meta.sheet == sheet].iloc[0]
        train_end = train_end_for(row["first"])
        target = build_sequences(daily[daily.sheet == sheet][["date", "pm25"]])
        if len(target) == 0:
            continue

        frames = source_frames(daily, train_end, row["first"])
        parts = {s: build_sequences(f) for s, f in frames.items()}
        parts = {s: d for s, d in parts.items() if len(d) > cfg.MIN_SOURCE_ROWS}
        if len(parts) < cfg.MIN_SOURCES:
            continue

        X_all = np.concatenate([d.X for d in parts.values()])
        y_all = np.concatenate([d.y for d in parts.values()])
        if len(X_all) > cfg.LSTM_MAX_SOURCE_SEQ:
            pick = np.random.default_rng(cfg.SEED).choice(
                len(X_all), cfg.LSTM_MAX_SOURCE_SEQ, replace=False
            )
            pick.sort()
            X_all, y_all = X_all[pick], y_all[pick]

        key = str(train_end.date())
        if key not in cache:
            shape = X_all.shape
            scaler = StandardScaler().fit(X_all.reshape(-1, shape[2]))
            log(f"[{n}/{len(targets)}] training source LSTM train_end={key} n={len(X_all)}")
            scaled = scaler.transform(X_all.reshape(-1, shape[2])).reshape(shape)
            cache[key] = (LSTMBackbone().fit_sources(scaled, y_all), scaler)
        lstm, scaler = cache[key]
        ridge = RidgeBackbone().fit_sources(X_all, y_all)

        # per-station relative (top-10%) threshold, over the target's full series,
        # so the LSTM is scored on the extrel slice and event-warning skill too.
        threshold = float(np.nanpercentile(target.y, cfg.EXTREME_PCT))

        for k_adapt in cfg.K_LIST:
            test_start = row["first"] + timedelta(days=k_adapt)
            index = build_common_index(target, not_before=test_start)
            if len(index) == 0:
                continue
            X_test, y_test = align(target, index)
            shape = X_test.shape
            X_scaled = scaler.transform(X_test.reshape(-1, shape[2])).reshape(shape)

            predictions = {
                "persistence": Prediction(X_test[:, -1, 0], index),
                "ridge": Prediction(ridge.predict(X_test), index),
                "lstm": Prediction(lstm.predict(X_scaled), index),
            }
            scores = score(y_test, predictions, index, relative_threshold=threshold)
            rows.append({
                "target": sheet, "name": row["name"], "K": k_adapt,
                "n_test": len(index), **scores,
            })

            # event-warning skill for all three, on the same common index
            if k_adapt == cfg.K_ADAPT_DEFAULT and (y_test >= threshold).sum() >= cfg.MIN_SLICE_DAYS:
                for name, pred in predictions.items():
                    p, r, f, n_ev = _detection(y_test, pred.values, threshold)
                    det.append({"target": sheet, "model": name, "precision": p,
                                "recall": r, "f1": f, "n_events": n_ev})

    frame = pd.DataFrame(rows)
    path = out_dir / "lstm.csv"
    frame.to_csv(path, index=False)
    det_frame = pd.DataFrame(det)
    det_frame.to_csv(out_dir / "lstm_detection.csv", index=False)
    for k_adapt in sorted(frame.K.unique()):
        subset = frame[frame.K == k_adapt]
        log(f"\n--- K={k_adapt} (n={len(subset)}) ---")
        log(summarise([
            compare(subset.rmse_all_persistence, subset.rmse_all_ridge, label="ridge"),
            compare(subset.rmse_all_persistence, subset.rmse_all_lstm, label="lstm"),
            compare(subset.rmse_all_lstm, subset.rmse_all_ridge, label="ridge vs lstm"),
            compare(subset.rmse_extrel_persistence, subset.rmse_extrel_lstm, label="lstm top10%"),
        ]))
    log(f"\nEVENT-WARNING SKILL (K={cfg.K_ADAPT_DEFAULT}, median over "
        f"{det_frame.target.nunique()} targets)")
    log(f"{'model':<14}{'precision':>11}{'recall':>9}{'F1':>8}")
    for name in ("persistence", "ridge", "lstm"):
        s = det_frame[det_frame.model == name]
        log(f"{name:<14}{s.precision.median():>11.2f}{s.recall.median():>9.2f}{s.f1.median():>8.2f}")
    log(f"\nsaved -> {path}, {out_dir / 'lstm_detection.csv'}")
    return frame


LOCAL_LAG1_ALPHA = 10.0
LOCAL_RIDGE_HI_ALPHA = 1000.0


def _adaptation_rows(design, first, test_start):
    """Rows of a target design whose predicted day falls in the adaptation window."""
    mask = (design.index >= first) & (design.index < test_start)
    return design.X[mask], design.y[mask]


def _lstm_representative(out_dir):
    """Per-(target,K) LSTM RMSE for the representative config h64_d0.2, from the
    committed sweep grid. Scored on the identical common index (build_sequences
    shares build_design's validity rule), so it merges by (target, K)."""
    for candidate in (out_dir / "lstm_sweep.csv", cfg.OUTPUT_DIR / "lstm_sweep.csv"):
        if candidate.exists():
            g = pd.read_csv(candidate)
            rep = g[(g.hidden == 64) & (g.dropout == 0.2)]
            return {(r.target, int(r.K)): r.lstm for _, r in rep.iterrows()}
    return {}


def stage_baselines(out_dir):
    """Target-only local models: answers "why not just train on the target's data?"

    Every model here is trained ONLY on the target's K-day adaptation window and
    scored on the same common index as every other method. The PRIMARY local model
    is a fair one (lag-1 feature, ridge alpha=10). Two SENSITIVITY variants
    (adaptation-window mean; ridge alpha=1000) bracket it. Two are reported as
    FINDINGS, not baselines: unregularised ridge (alpha=1) and RidgeCV both
    collapse when fit on ~16 adaptation samples with 18 features -- selecting or
    trusting alpha on so few points is unreliable.

    Features are standardised on the adaptation window only. A target with fewer
    than 2 adaptation sequences (station_20 at K=30, 0 sequences) is dropped at
    that K -- a local model cannot be trained there at all, which is part of the
    finding.
    """
    daily, meta, wide1, wide2 = _context()
    targets = select_targets(daily, meta)
    banner("BASELINES — target-only local models vs persistence, transfer, LSTM")
    lstm_ref = _lstm_representative(out_dir)

    rows = []
    for sheet in targets:
        row = meta[meta.sheet == sheet].iloc[0]
        first = row["first"]
        train_end = train_end_for(first)
        design = build_design(daily[daily.sheet == sheet][["date", "pm25"]])

        for k_adapt in cfg.K_LIST:
            test_start = first + timedelta(days=k_adapt)
            index = build_common_index(design, not_before=test_start)
            if len(index) == 0:
                continue
            X_test, y_test = align(design, index)
            X_adapt, y_adapt = _adaptation_rows(design, first, test_start)
            if len(y_adapt) < 2:
                log(f"  skip {sheet} K={k_adapt}: {len(y_adapt)} adaptation sequences "
                    f"(local model undefined)")
                continue

            Xtr, ytr = _pooled_source_design(daily, meta, train_end, first, 0, wide1, wide2)
            if Xtr is None:
                continue

            # local models: standardise on the adaptation window only
            col = slice(cfg.SEQ_L - 1, cfg.SEQ_L)  # the lag-1 column
            sc = StandardScaler().fit(X_adapt)
            Xa, Xt = sc.transform(X_adapt), sc.transform(X_test)
            sc1 = StandardScaler().fit(X_adapt[:, col])
            Xa1, Xt1 = sc1.transform(X_adapt[:, col]), sc1.transform(X_test[:, col])

            def fit_pred(model, xa, xt):
                return model.fit(xa, y_adapt).predict(xt)

            preds = {
                "persistence": Prediction(X_test[:, cfg.SEQ_L - 1], index),
                "pooled_transfer": Prediction(
                    RidgeBackbone().fit_sources(Xtr, ytr).predict(X_test), index),
                "local_lag1": Prediction(fit_pred(Ridge(alpha=LOCAL_LAG1_ALPHA), Xa1, Xt1), index),
                "local_mean": Prediction(np.full(len(index), float(y_adapt.mean())), index),
                "local_ridge_hi": Prediction(fit_pred(Ridge(alpha=LOCAL_RIDGE_HI_ALPHA), Xa, Xt), index),
                "local_ridge_unreg": Prediction(fit_pred(Ridge(alpha=1.0), Xa, Xt), index),
                "local_ridgecv": Prediction(
                    fit_pred(RidgeCV(alphas=np.logspace(-3, 3, 13)), Xa, Xt), index),
            }
            scored = score(y_test, preds, index)
            rec = {"target": sheet, "name": row["name"], "K": k_adapt,
                   "n_test": len(index), "n_adapt": len(y_adapt),
                   "lstm_repr": lstm_ref.get((sheet, k_adapt), np.nan)}
            rec.update({m: scored[f"rmse_all_{m}"] for m in preds})
            rec.update({f"mae_{m}": scored[f"mae_all_{m}"] for m in preds})
            rows.append(rec)

    frame = pd.DataFrame(rows)
    path = out_dir / "baselines.csv"
    frame.to_csv(path, index=False)

    ladder = ["local_ridge_unreg", "local_ridgecv", "local_mean", "local_ridge_hi",
              "local_lag1", "persistence", "lstm_repr", "pooled_transfer"]
    for k_adapt in sorted(frame.K.unique()):
        sub = frame[frame.K == k_adapt]
        log(f"\n--- K={k_adapt}  (n={len(sub)} targets with a usable adaptation window) ---")
        log(f"{'method':<20}{'median RMSE':>12}   role")
        role = {"local_ridge_unreg": "FINDING (alpha=1 collapses)",
                "local_ridgecv": "FINDING (CV on ~16 pts)",
                "local_mean": "sensitivity", "local_ridge_hi": "sensitivity",
                "local_lag1": "PRIMARY local", "persistence": "untrained baseline",
                "lstm_repr": "LSTM (repr h64_d0.2)", "pooled_transfer": "transfer"}
        for m in ladder:
            if m in sub and sub[m].notna().any():
                log(f"{m:<20}{sub[m].median():>12.2f}   {role[m]}")
        log("")
        # Paired comparisons, each with the method of interest as `b` so its
        # win-count is what prints. gain% (difference of medians) and dPair
        # (median paired difference) sit side by side -- when they disagree, the
        # loss distribution is skewed and dPair/wins/p are the ones to trust.
        log(summarise([
            compare(sub.persistence, sub.pooled_transfer, label="transfer vs pers"),
            compare(sub.persistence, sub.local_lag1, label="local vs pers"),
            compare(sub.local_lag1, sub.pooled_transfer, label="transfer vs local"),
        ]))
    log(f"\nsaved -> {path}")
    return frame


RIDGE_ALPHAS = (0.1, 1.0, 10.0, 100.0, 1000.0)


def _lstm_config(out_dir, hidden, dropout):
    """Per-(target,K) LSTM RMSE for one config, from the committed sweep grid."""
    for candidate in (out_dir / "lstm_sweep.csv", cfg.OUTPUT_DIR / "lstm_sweep.csv"):
        if candidate.exists():
            g = pd.read_csv(candidate)
            sub = g[(g.hidden == hidden) & (g.dropout == dropout)]
            return {(r.target, int(r.K)): r.lstm for _, r in sub.iterrows()}
    return {}


def stage_ridge_alpha(out_dir):
    """Ridge-penalty sensitivity of pooled transfer.

    The LSTM got a 9-config sweep; the ridge penalty was fixed at alpha=1.0 and
    never tuned. This closes that asymmetry: it runs pooled transfer at
    alpha in {0.1, 1, 10, 100, 1000} on the same common index, and checks whether
    "ridge beats persistence 12/12" and "ridge beats the LSTM" hold at every alpha
    -- against both the representative (h64_d0.2) and best-on-test (h32_d0.3) LSTM
    configs. Does NOT change the production alpha; if a different one is clearly
    better that is flagged for a human decision.
    """
    daily, meta, wide1, wide2 = _context()
    targets = select_targets(daily, meta)
    banner("RIDGE ALPHA — sensitivity of pooled transfer to the ridge penalty")
    lstm_repr = _lstm_config(out_dir, 64, 0.2)
    lstm_oracle = _lstm_config(out_dir, 32, 0.3)

    rows = []
    for sheet in targets:
        row = meta[meta.sheet == sheet].iloc[0]
        train_end = train_end_for(row["first"])
        design = build_design(daily[daily.sheet == sheet][["date", "pm25"]])
        Xtr, ytr = _pooled_source_design(daily, meta, train_end, row["first"], 0, wide1, wide2)
        if Xtr is None:
            continue
        for k_adapt in cfg.K_LIST:
            test_start = row["first"] + timedelta(days=k_adapt)
            index = build_common_index(design, not_before=test_start)
            if len(index) == 0:
                continue
            X_test, y_test = align(design, index)
            preds = {"persistence": Prediction(X_test[:, cfg.SEQ_L - 1], index)}
            for a in RIDGE_ALPHAS:
                preds[f"ridge_a{a}"] = Prediction(
                    RidgeBackbone(alpha=a).fit_sources(Xtr, ytr).predict(X_test), index)
            scored = score(y_test, preds, index)
            rec = {"target": sheet, "name": row["name"], "K": k_adapt, "n_test": len(index),
                   "persistence": scored["rmse_all_persistence"],
                   "lstm_repr": lstm_repr.get((sheet, k_adapt), np.nan),
                   "lstm_oracle": lstm_oracle.get((sheet, k_adapt), np.nan)}
            rec.update({f"ridge_a{a}": scored[f"rmse_all_ridge_a{a}"] for a in RIDGE_ALPHAS})
            rows.append(rec)

    frame = pd.DataFrame(rows)
    frame.to_csv(out_dir / "ridge_alpha_sensitivity.csv", index=False)

    for k_adapt in sorted(frame.K.unique()):
        sub = frame[frame.K == k_adapt]
        log(f"\n--- K={k_adapt}  (n={len(sub)}) ---")
        log(f"{'alpha':<8}{'medRMSE':>9}{'vs persistence':>18}"
            f"{'vs LSTM repr':>18}{'vs LSTM oracle':>18}")
        for a in RIDGE_ALPHAS:
            col = f"ridge_a{a}"
            cp = compare(sub.persistence, sub[col])   # b=ridge
            cr = compare(sub.lstm_repr, sub[col])
            co = compare(sub.lstm_oracle, sub[col])
            mark = " <-prod" if a == cfg.RIDGE_ALPHA else ""
            log(f"{a:<8}{sub[col].median():>9.3f}"
                f"{f'{cp.b_wins}/{cp.n},p={cp.p_value:.3f}':>18}"
                f"{f'{cr.b_wins}/{cr.n},p={cr.p_value:.3f}':>18}"
                f"{f'{co.b_wins}/{co.n},p={co.p_value:.3f}':>18}{mark}")
        best = min(RIDGE_ALPHAS, key=lambda a: sub[f"ridge_a{a}"].median())
        prod_med = sub[f"ridge_a{cfg.RIDGE_ALPHA}"].median()
        best_med = sub[f"ridge_a{best}"].median()
        log(f"  best-median alpha: {best} ({best_med:.3f}) vs production "
            f"{cfg.RIDGE_ALPHA} ({prod_med:.3f}); gap {prod_med - best_med:+.3f} RMSE")
    log(f"\nsaved -> {out_dir / 'ridge_alpha_sensitivity.csv'}")
    return frame


STAGES = {
    "audit": stage_audit,
    "selection": stage_selection,
    "baselines": stage_baselines,
    "ridge_alpha": stage_ridge_alpha,
    "ablation": stage_ablation,
    "extreme": stage_extreme,
    "lstm": stage_lstm,
}
